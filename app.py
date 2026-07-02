from __future__ import annotations

import csv
import io
import json
import os
import uuid
from datetime import date, datetime
from http.cookies import SimpleCookie
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qsl, parse_qs, urlencode, urlparse, urlunparse

try:
    from openpyxl import Workbook, load_workbook
except ImportError:  # pragma: no cover - import guard for clearer startup errors
    Workbook = None
    load_workbook = None

try:
    import psycopg
    from psycopg.rows import dict_row
except ImportError:  # pragma: no cover - local SQLite mode does not need psycopg
    psycopg = None
    dict_row = None

import sqlite3


ROOT = Path(__file__).resolve().parent
DB_PATH = Path(os.environ.get("DATABASE_PATH", ROOT / "data" / "project_tracker.db"))
DATABASE_URL = os.environ.get("DATABASE_URL")
USING_POSTGRES = bool(DATABASE_URL)
REQUIRE_DATABASE_URL = os.environ.get("REQUIRE_DATABASE_URL", "").lower() in {"1", "true", "yes"}
STATIC_DIR = ROOT / "static"
APP_PASSWORD = os.environ.get("TRACKER_PASSWORD", "change-me")
SESSIONS: set[str] = set()

STATUSES = [
    "Planned",
    "Survey",
    "Procurement",
    "Implementation",
    "Testing",
    "Delivered",
    "Confirmed",
    "On Hold",
    "Cancelled",
]

PROJECT_FIELDS = [
    "sn",
    "customer_name",
    "site_name",
    "location",
    "region",
    "capacity",
    "bandwidth",
    "planned_adss_distance_m",
    "planned_drop_distance_m",
    "service_type",
    "cpe",
    "confirmation_date",
    "start_date",
    "target_completion_date",
    "completion_date",
    "status",
    "remarks",
    "currency",
    "mrc",
    "nrc",
    "poles_9m",
    "poles_11m",
    "pole_cost",
    "transportation_cost",
    "labour_cost",
    "total_cost",
    "labour",
    "trench",
    "parent_project_id",
    "archived",
]

NUMERIC_FIELDS = {
    "planned_adss_distance_m",
    "planned_drop_distance_m",
    "mrc",
    "nrc",
    "poles_9m",
    "poles_11m",
    "pole_cost",
    "transportation_cost",
    "labour_cost",
    "total_cost",
    "labour",
    "trench",
}

DATE_FIELDS = {
    "confirmation_date",
    "start_date",
    "target_completion_date",
    "completion_date",
}

HEADER_ALIASES = {
    "project id": "project_id",
    "column1": "sn",
    "sn": "sn",
    "customer name": "customer_name",
    "customer": "customer_name",
    "site name": "site_name",
    "site": "site_name",
    "location": "location",
    "region": "region",
    "capacity": "capacity",
    "bandwidth": "bandwidth",
    "planned adss distance (m)": "planned_adss_distance_m",
    "planned adss distance": "planned_adss_distance_m",
    "adss": "planned_adss_distance_m",
    "planned drop distance (m)": "planned_drop_distance_m",
    "planned drop distance": "planned_drop_distance_m",
    "drop": "planned_drop_distance_m",
    "service type": "service_type",
    "product": "service_type",
    "cpe": "cpe",
    "confirmation date": "confirmation_date",
    "order received date": "confirmation_date",
    "date submitted": "confirmation_date",
    "start date": "start_date",
    "billing start date": "start_date",
    "target completion date": "target_completion_date",
    "estimated completion": "target_completion_date",
    "completion date": "completion_date",
    "delivery date": "completion_date",
    "date completed": "completion_date",
    "status": "status",
    "remarks": "remarks",
    "comments": "remarks",
    "currency": "currency",
    "mrc": "mrc",
    "nrc": "nrc",
    "9m": "poles_9m",
    "9m poles": "poles_9m",
    "11m": "poles_11m",
    "11m poles": "poles_11m",
    "pole cost": "pole_cost",
    "transportation": "transportation_cost",
    "transportation cost": "transportation_cost",
    "labour (chs)": "labour_cost",
    "labor (chs)": "labour_cost",
    "labour cost": "labour_cost",
    "labor cost": "labour_cost",
    "total cost": "total_cost",
    "labour": "labour",
    "labor": "labour",
    "trench": "trench",
    "parent project id": "parent_project_id",
    "archived": "archived",
}


def sql(statement: str) -> str:
    return statement.replace("?", "%s") if USING_POSTGRES else statement


def postgres_url() -> str:
    parsed = urlparse(DATABASE_URL)
    ignored_params = {"pgbouncer", "connection_limit"}
    query = [(key, value) for key, value in parse_qsl(parsed.query, keep_blank_values=True) if key not in ignored_params]
    return urlunparse(parsed._replace(query=urlencode(query)))


def connect():
    if USING_POSTGRES:
        if psycopg is None:
            raise RuntimeError("psycopg is required when DATABASE_URL is set")
        return psycopg.connect(postgres_url(), row_factory=dict_row)
    DB_PATH.parent.mkdir(exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db() -> None:
    with connect() as conn:
        if USING_POSTGRES:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS projects (
                    id SERIAL PRIMARY KEY,
                    project_id TEXT NOT NULL UNIQUE,
                    sn TEXT,
                    customer_name TEXT NOT NULL,
                    site_name TEXT NOT NULL,
                    location TEXT,
                    region TEXT,
                    capacity TEXT,
                    bandwidth DOUBLE PRECISION,
                    planned_adss_distance_m DOUBLE PRECISION,
                    planned_drop_distance_m DOUBLE PRECISION,
                    service_type TEXT,
                    cpe TEXT,
                    confirmation_date TEXT,
                    start_date TEXT,
                    target_completion_date TEXT,
                    completion_date TEXT,
                    status TEXT NOT NULL DEFAULT 'Planned',
                    remarks TEXT,
                    currency TEXT DEFAULT 'GHS',
                    mrc DOUBLE PRECISION DEFAULT 0,
                    nrc DOUBLE PRECISION DEFAULT 0,
                    poles_9m DOUBLE PRECISION DEFAULT 0,
                    poles_11m DOUBLE PRECISION DEFAULT 0,
                    pole_cost DOUBLE PRECISION DEFAULT 0,
                    transportation_cost DOUBLE PRECISION DEFAULT 0,
                    labour_cost DOUBLE PRECISION DEFAULT 0,
                    total_cost DOUBLE PRECISION DEFAULT 0,
                    labour DOUBLE PRECISION DEFAULT 0,
                    trench DOUBLE PRECISION DEFAULT 0,
                    parent_project_id TEXT,
                    archived BOOLEAN DEFAULT FALSE,
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL
                )
                """
            )
        else:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS projects (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    project_id TEXT NOT NULL UNIQUE,
                    sn TEXT,
                    customer_name TEXT NOT NULL,
                    site_name TEXT NOT NULL,
                    location TEXT,
                    region TEXT,
                    capacity TEXT,
                    bandwidth REAL,
                    planned_adss_distance_m REAL,
                    planned_drop_distance_m REAL,
                    service_type TEXT,
                    cpe TEXT,
                    confirmation_date TEXT,
                    start_date TEXT,
                    target_completion_date TEXT,
                    completion_date TEXT,
                    status TEXT NOT NULL DEFAULT 'Planned',
                    remarks TEXT,
                    currency TEXT DEFAULT 'GHS',
                    mrc REAL DEFAULT 0,
                    nrc REAL DEFAULT 0,
                    poles_9m REAL DEFAULT 0,
                    poles_11m REAL DEFAULT 0,
                    pole_cost REAL DEFAULT 0,
                    transportation_cost REAL DEFAULT 0,
                    labour_cost REAL DEFAULT 0,
                    total_cost REAL DEFAULT 0,
                    labour REAL DEFAULT 0,
                    trench REAL DEFAULT 0,
                    parent_project_id TEXT,
                    archived INTEGER DEFAULT 0,
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL
                )
                """
            )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_projects_status ON projects(status)"
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_projects_region ON projects(region)"
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_projects_customer ON projects(customer_name)"
        )
        ensure_project_columns(conn)


def ensure_project_columns(conn) -> None:
    columns = {
        "currency": "TEXT DEFAULT 'GHS'",
        "pole_cost": "DOUBLE PRECISION DEFAULT 0" if USING_POSTGRES else "REAL DEFAULT 0",
        "transportation_cost": "DOUBLE PRECISION DEFAULT 0" if USING_POSTGRES else "REAL DEFAULT 0",
        "labour_cost": "DOUBLE PRECISION DEFAULT 0" if USING_POSTGRES else "REAL DEFAULT 0",
        "total_cost": "DOUBLE PRECISION DEFAULT 0" if USING_POSTGRES else "REAL DEFAULT 0",
    }
    if USING_POSTGRES:
        existing = {
            row["column_name"]
            for row in conn.execute(
                """
                SELECT column_name
                FROM information_schema.columns
                WHERE table_name = 'projects'
                """
            ).fetchall()
        }
    else:
        existing = {row["name"] for row in conn.execute("PRAGMA table_info(projects)").fetchall()}
    for name, definition in columns.items():
        if name not in existing:
            conn.execute(f"ALTER TABLE projects ADD COLUMN {name} {definition}")


def today() -> date:
    return date.today()


def parse_date(value) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def iso_date(value) -> str | None:
    parsed = parse_date(value)
    return parsed.isoformat() if parsed else None


def normalize_number(value):
    if value in (None, ""):
        return 0
    if isinstance(value, (int, float)):
        return value
    cleaned = str(value).replace(",", "").strip()
    try:
        return float(cleaned)
    except ValueError:
        return 0


def normalize_bool(value) -> int:
    if isinstance(value, bool):
        return int(value)
    return 1 if str(value).strip().lower() in {"1", "yes", "true", "archived"} else 0


def normalize_archived(value):
    archived = bool(normalize_bool(value))
    return archived if USING_POSTGRES else int(archived)


def make_project_id() -> str:
    return f"PRJ-{datetime.now().strftime('%Y%m%d')}-{uuid.uuid4().hex[:5].upper()}"


def normalize_project(payload: dict, existing_project_id: str | None = None) -> dict:
    data = {}
    for field in PROJECT_FIELDS:
        value = payload.get(field)
        if field in NUMERIC_FIELDS or field == "bandwidth":
            data[field] = normalize_number(value)
        elif field in DATE_FIELDS:
            data[field] = iso_date(value)
        elif field == "archived":
            data[field] = normalize_archived(value)
        else:
            data[field] = str(value).strip() if value not in (None, "") else None

    data["customer_name"] = data["customer_name"] or "Unnamed Customer"
    data["site_name"] = data["site_name"] or "Unnamed Site"
    data["currency"] = data["currency"] or "GHS"
    data["status"] = data["status"] if data["status"] in STATUSES else "Planned"
    data["project_id"] = payload.get("project_id") or existing_project_id or make_project_id()
    apply_cost_calculations(data)
    return data


def apply_cost_calculations(data: dict) -> None:
    poles_9m = normalize_number(data.get("poles_9m"))
    poles_11m = normalize_number(data.get("poles_11m"))
    adss = normalize_number(data.get("planned_adss_distance_m"))
    drop = normalize_number(data.get("planned_drop_distance_m"))
    data["pole_cost"] = (2000 * poles_9m) + (3000 * poles_11m)
    data["transportation_cost"] = 500 * poles_9m + 500 * poles_11m
    data["labour_cost"] = (2 * adss) + (2 * drop)
    data["total_cost"] = data["pole_cost"] + data["transportation_cost"] + data["labour_cost"]


def row_to_project(row: sqlite3.Row) -> dict:
    project = dict(row)
    apply_cost_calculations(project)
    target = parse_date(project.get("target_completion_date"))
    confirmation = parse_date(project.get("confirmation_date"))
    completion = parse_date(project.get("completion_date"))
    end = completion or today()
    project["is_delayed"] = bool(target and target < today() and project["status"] not in {"Delivered", "Confirmed", "Cancelled"})
    project["aging_days"] = max((end - confirmation).days, 0) if confirmation else None
    project["aging_bucket"] = aging_bucket(project["aging_days"])
    project["archived"] = bool(project["archived"])
    return project


def aging_bucket(days) -> str:
    if days is None:
        return "Unconfirmed"
    if days <= 30:
        return "0-30"
    if days <= 60:
        return "31-60"
    if days <= 90:
        return "61-90"
    if days <= 120:
        return "91-120"
    return "120+"


def get_projects(query: dict) -> list[dict]:
    clauses = []
    params = []

    if query.get("status"):
        clauses.append("status = ?")
        params.append(query["status"])
    if query.get("region"):
        clauses.append("region = ?")
        params.append(query["region"])
    if query.get("service_type"):
        clauses.append("service_type = ?")
        params.append(query["service_type"])
    if query.get("customer"):
        clauses.append("customer_name = ?")
        params.append(query["customer"])
    if query.get("archived") in {"0", "1"}:
        clauses.append("archived = ?")
        params.append(bool(int(query["archived"])) if USING_POSTGRES else int(query["archived"]))
    if query.get("q"):
        clauses.append("(site_name LIKE ? OR customer_name LIKE ? OR remarks LIKE ?)")
        search = f"%{query['q']}%"
        params.extend([search, search, search])
    if query.get("from"):
        clauses.append("confirmation_date >= ?")
        params.append(query["from"])
    if query.get("to"):
        clauses.append("confirmation_date <= ?")
        params.append(query["to"])
    if query.get("min_bandwidth"):
        clauses.append("bandwidth >= ?")
        params.append(normalize_number(query["min_bandwidth"]))
    if query.get("max_bandwidth"):
        clauses.append("bandwidth <= ?")
        params.append(normalize_number(query["max_bandwidth"]))

    where = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    with connect() as conn:
        rows = conn.execute(
            sql(f"SELECT * FROM projects {where} ORDER BY archived ASC, confirmation_date DESC, id DESC"),
            params,
        ).fetchall()
    return [row_to_project(row) for row in rows]


def save_project(payload: dict, project_pk: int | None = None) -> dict:
    now = datetime.utcnow().isoformat(timespec="seconds")
    existing = None
    if project_pk:
        with connect() as conn:
            existing = conn.execute(sql("SELECT * FROM projects WHERE id = ?"), (project_pk,)).fetchone()
        if not existing:
            raise KeyError("Project not found")

    data = normalize_project(payload, existing["project_id"] if existing else None)
    data["updated_at"] = now

    with connect() as conn:
        if project_pk:
            assignments = ", ".join([f"{field} = ?" for field in ["project_id"] + PROJECT_FIELDS] + ["updated_at = ?"])
            values = [data["project_id"]] + [data[field] for field in PROJECT_FIELDS] + [data["updated_at"], project_pk]
            conn.execute(sql(f"UPDATE projects SET {assignments} WHERE id = ?"), values)
        else:
            data["created_at"] = now
            fields = ["project_id"] + PROJECT_FIELDS + ["created_at", "updated_at"]
            placeholders = ", ".join(["?"] * len(fields))
            conn.execute(
                sql(f"INSERT INTO projects ({', '.join(fields)}) VALUES ({placeholders})"),
                [data[field] for field in fields],
            )
        row = conn.execute(sql("SELECT * FROM projects WHERE project_id = ?"), (data["project_id"],)).fetchone()
    return row_to_project(row)


def delete_project(project_pk: int) -> None:
    with connect() as conn:
        conn.execute(sql("DELETE FROM projects WHERE id = ?"), (project_pk,))


def summary() -> dict:
    projects = get_projects({})
    active = [p for p in projects if not p["archived"]]
    completed = [p for p in projects if p["status"] in {"Delivered", "Confirmed"}]
    delayed = [p for p in active if p["is_delayed"]]
    completed_with_age = [p["aging_days"] for p in completed if p["aging_days"] is not None]

    def by_field(field: str, value_field: str | None = None):
        grouped = {}
        for p in projects:
            key = p.get(field) or "Unassigned"
            if value_field:
                grouped[key] = grouped.get(key, 0) + float(p.get(value_field) or 0)
            else:
                grouped[key] = grouped.get(key, 0) + 1
        return grouped

    aging = {}
    for p in active:
        bucket = p["aging_bucket"]
        aging[bucket] = aging.get(bucket, 0) + 1

    today_date = today()
    week_start = today_date.fromordinal(today_date.toordinal() - today_date.weekday())
    recent_today = []
    recent_week = []
    for p in projects:
        created = parse_date((p.get("created_at") or "").split("T")[0])
        if created == today_date:
            recent_today.append(p)
        if created and created >= week_start:
            recent_week.append(p)

    return {
        "total_projects": len(projects),
        "active_projects": len(active),
        "completed_projects": len(completed),
        "in_progress_projects": len([p for p in active if p["status"] not in {"Delivered", "Confirmed", "Cancelled"}]),
        "delayed_projects": len(delayed),
        "average_completion_days": round(sum(completed_with_age) / len(completed_with_age), 1) if completed_with_age else 0,
        "mrc": sum(float(p.get("mrc") or 0) for p in projects),
        "nrc": sum(float(p.get("nrc") or 0) for p in projects),
        "materials": {
            "planned_adss_distance_m": sum(float(p.get("planned_adss_distance_m") or 0) for p in projects),
            "planned_drop_distance_m": sum(float(p.get("planned_drop_distance_m") or 0) for p in projects),
            "poles_9m": sum(float(p.get("poles_9m") or 0) for p in projects),
            "poles_11m": sum(float(p.get("poles_11m") or 0) for p in projects),
            "labour": sum(float(p.get("labour") or 0) for p in projects),
            "trench": sum(float(p.get("trench") or 0) for p in projects),
        },
        "costs": {
            "pole_cost": sum(float(p.get("pole_cost") or 0) for p in projects),
            "transportation_cost": sum(float(p.get("transportation_cost") or 0) for p in projects),
            "labour_cost": sum(float(p.get("labour_cost") or 0) for p in projects),
            "total_cost": sum(float(p.get("total_cost") or 0) for p in projects),
        },
        "recent": {
            "today": len(recent_today),
            "week": len(recent_week),
            "items": recent_week[:10],
        },
        "aging": aging,
        "by_status": by_field("status"),
        "revenue_by_region": by_field("region", "mrc"),
        "revenue_by_customer": by_field("customer_name", "mrc"),
        "by_service_type": by_field("service_type"),
    }


def distinct_options() -> dict:
    with connect() as conn:
        options = {}
        for field in ["customer_name", "region", "service_type", "status"]:
            rows = conn.execute(
                f"SELECT DISTINCT {field} value FROM projects WHERE {field} IS NOT NULL AND {field} != '' ORDER BY {field}"
            ).fetchall()
            options[field] = [row["value"] for row in rows]
    options["statuses"] = STATUSES
    return options


def export_workbook(projects: list[dict]) -> bytes:
    if Workbook is None:
        raise RuntimeError("openpyxl is required for Excel export")
    wb = Workbook()
    ws = wb.active
    ws.title = "Projects"
    headers = [
        "Project ID",
        "SN",
        "Customer Name",
        "Site Name",
        "Location",
        "Region",
        "Currency",
        "MRC",
        "NRC",
        "Capacity",
        "Bandwidth",
        "Planned ADSS Distance (m)",
        "Planned Drop Distance (m)",
        "Service Type",
        "CPE",
        "Confirmation Date",
        "Start Date",
        "Target Completion Date",
        "Completion Date",
        "Aging Days",
        "Aging Bucket",
        "Delayed",
        "Status",
        "Remarks",
        "9m Poles",
        "11m Poles",
        "Pole Cost",
        "Transportation",
        "Labour Cost",
        "Total Cost",
        "Labour",
        "Trench",
        "Parent Project ID",
        "Archived",
    ]
    ws.append(headers)
    for p in projects:
        ws.append([
            p.get("project_id"),
            p.get("sn"),
            p.get("customer_name"),
            p.get("site_name"),
            p.get("location"),
            p.get("region"),
            p.get("currency"),
            p.get("mrc"),
            p.get("nrc"),
            p.get("capacity"),
            p.get("bandwidth"),
            p.get("planned_adss_distance_m"),
            p.get("planned_drop_distance_m"),
            p.get("service_type"),
            p.get("cpe"),
            p.get("confirmation_date"),
            p.get("start_date"),
            p.get("target_completion_date"),
            p.get("completion_date"),
            p.get("aging_days"),
            p.get("aging_bucket"),
            "Yes" if p.get("is_delayed") else "No",
            p.get("status"),
            p.get("remarks"),
            p.get("poles_9m"),
            p.get("poles_11m"),
            p.get("pole_cost"),
            p.get("transportation_cost"),
            p.get("labour_cost"),
            p.get("total_cost"),
            p.get("labour"),
            p.get("trench"),
            p.get("parent_project_id"),
            "Yes" if p.get("archived") else "No",
        ])
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_length + 2, 12), 34)
    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


def import_workbook(file_bytes: bytes) -> dict:
    if load_workbook is None:
        raise RuntimeError("openpyxl is required for Excel import")
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb["2026 WIP Tracker"] if "2026 WIP Tracker" in wb.sheetnames else max(wb.worksheets, key=lambda sheet: sheet.max_row * sheet.max_column)
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"imported": 0, "updated": 0, "errors": []}
    headers = []
    for index, header in enumerate(rows[0], start=1):
        key = str(header or "").strip().lower()
        mapped = HEADER_ALIASES.get(key)
        if not mapped and index == 2:
            mapped = "customer_name"
        headers.append(mapped)
    imported = 0
    updated = 0
    errors = []
    with connect() as conn:
        for index, row in enumerate(rows[1:], start=2):
            payload = {}
            for header, value in zip(headers, row):
                if header:
                    payload[header] = value
            if not any(payload.values()):
                continue
            try:
                project_id = payload.get("project_id")
                existing = None
                if project_id:
                    existing = conn.execute(sql("SELECT id FROM projects WHERE project_id = ?"), (project_id,)).fetchone()
                saved = save_project(payload, existing["id"] if existing else None)
                updated += 1 if existing else 0
                imported += 0 if existing else 1
            except Exception as exc:  # pragma: no cover - user data path
                errors.append(f"Row {index}: {exc}")
    return {"imported": imported, "updated": updated, "errors": errors}


def json_response(handler: BaseHTTPRequestHandler, payload, status=HTTPStatus.OK) -> None:
    body = json.dumps(payload, default=str).encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", "application/json")
    handler.send_header("Content-Length", str(len(body)))
    handler.end_headers()
    handler.wfile.write(body)


def plain_response(handler: BaseHTTPRequestHandler, text: str, status=HTTPStatus.OK) -> None:
    body = text.encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", "text/html; charset=utf-8")
    handler.send_header("Content-Length", str(len(body)))
    handler.end_headers()
    handler.wfile.write(body)


class TrackerHandler(BaseHTTPRequestHandler):
    def do_GET(self):
        parsed = urlparse(self.path)
        path = parsed.path
        query = {k: v[0] for k, v in parse_qs(parsed.query).items()}
        if path == "/api/health":
            return json_response(self, {"ok": True})
        if path == "/login":
            return self.serve_static("/login.html", require_auth=False)
        if path == "/styles.css":
            return self.serve_static(path, require_auth=False)
        if not self.is_authenticated():
            if path.startswith("/api/"):
                return json_response(self, {"error": "Authentication required"}, HTTPStatus.UNAUTHORIZED)
            self.send_response(HTTPStatus.FOUND)
            self.send_header("Location", "/login")
            self.end_headers()
            return
        if path == "/api/projects":
            return json_response(self, {"projects": get_projects(query)})
        if path == "/api/summary":
            return json_response(self, summary())
        if path == "/api/options":
            return json_response(self, distinct_options())
        if path == "/api/export":
            projects = get_projects(query)
            data = export_workbook(projects)
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", "attachment; filename=project-tracker-export.xlsx")
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            self.wfile.write(data)
            return
        return self.serve_static(path)

    def do_POST(self):
        parsed = urlparse(self.path)
        if parsed.path == "/api/login":
            payload = self.read_json()
            if payload.get("password") == APP_PASSWORD:
                token = uuid.uuid4().hex
                SESSIONS.add(token)
                body = json.dumps({"ok": True}).encode("utf-8")
                self.send_response(HTTPStatus.OK)
                self.send_header("Content-Type", "application/json")
                self.send_header("Set-Cookie", f"tracker_session={token}; Path=/; HttpOnly; SameSite=Lax")
                self.send_header("Content-Length", str(len(body)))
                self.end_headers()
                self.wfile.write(body)
                return
            return json_response(self, {"error": "Invalid password"}, HTTPStatus.UNAUTHORIZED)
        if parsed.path == "/api/logout":
            token = self.session_token()
            if token:
                SESSIONS.discard(token)
            body = json.dumps({"ok": True}).encode("utf-8")
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "application/json")
            self.send_header("Set-Cookie", "tracker_session=; Path=/; Max-Age=0; HttpOnly; SameSite=Lax")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
            return
        if not self.is_authenticated():
            return json_response(self, {"error": "Authentication required"}, HTTPStatus.UNAUTHORIZED)
        if parsed.path == "/api/projects":
            payload = self.read_json()
            try:
                return json_response(self, save_project(payload), HTTPStatus.CREATED)
            except Exception as exc:
                return json_response(self, {"error": str(exc)}, HTTPStatus.BAD_REQUEST)
        if parsed.path == "/api/import":
            length = int(self.headers.get("Content-Length", 0))
            file_bytes = self.rfile.read(length)
            try:
                return json_response(self, import_workbook(file_bytes))
            except Exception as exc:
                return json_response(self, {"error": str(exc)}, HTTPStatus.BAD_REQUEST)
        return json_response(self, {"error": "Not found"}, HTTPStatus.NOT_FOUND)

    def do_PUT(self):
        parsed = urlparse(self.path)
        if not self.is_authenticated():
            return json_response(self, {"error": "Authentication required"}, HTTPStatus.UNAUTHORIZED)
        if parsed.path.startswith("/api/projects/"):
            project_pk = int(parsed.path.rsplit("/", 1)[-1])
            try:
                return json_response(self, save_project(self.read_json(), project_pk))
            except KeyError:
                return json_response(self, {"error": "Project not found"}, HTTPStatus.NOT_FOUND)
            except Exception as exc:
                return json_response(self, {"error": str(exc)}, HTTPStatus.BAD_REQUEST)
        return json_response(self, {"error": "Not found"}, HTTPStatus.NOT_FOUND)

    def do_DELETE(self):
        parsed = urlparse(self.path)
        if not self.is_authenticated():
            return json_response(self, {"error": "Authentication required"}, HTTPStatus.UNAUTHORIZED)
        if parsed.path.startswith("/api/projects/"):
            delete_project(int(parsed.path.rsplit("/", 1)[-1]))
            self.send_response(HTTPStatus.NO_CONTENT)
            self.end_headers()
            return
        return json_response(self, {"error": "Not found"}, HTTPStatus.NOT_FOUND)

    def read_json(self) -> dict:
        length = int(self.headers.get("Content-Length", 0))
        body = self.rfile.read(length).decode("utf-8")
        return json.loads(body or "{}")

    def session_token(self) -> str | None:
        cookie = SimpleCookie(self.headers.get("Cookie"))
        morsel = cookie.get("tracker_session")
        return morsel.value if morsel else None

    def is_authenticated(self) -> bool:
        return self.session_token() in SESSIONS

    def serve_static(self, path: str, require_auth: bool = True):
        if require_auth and not self.is_authenticated():
            self.send_response(HTTPStatus.FOUND)
            self.send_header("Location", "/login")
            self.end_headers()
            return
        if path == "/":
            path = "/index.html"
        target = (STATIC_DIR / path.lstrip("/")).resolve()
        if not str(target).startswith(str(STATIC_DIR.resolve())) or not target.exists():
            return json_response(self, {"error": "Not found"}, HTTPStatus.NOT_FOUND)
        content_types = {
            ".html": "text/html; charset=utf-8",
            ".css": "text/css; charset=utf-8",
            ".js": "application/javascript; charset=utf-8",
            ".svg": "image/svg+xml",
        }
        body = target.read_bytes()
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", content_types.get(target.suffix, "application/octet-stream"))
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def log_message(self, format, *args):  # noqa: A002
        print(f"{self.address_string()} - {format % args}")


def seed_if_empty() -> None:
    with connect() as conn:
        count = conn.execute("SELECT COUNT(*) total FROM projects").fetchone()["total"]
    if count:
        return
    samples = [
        {
            "customer_name": "Acme Fibre",
            "site_name": "Airport Branch",
            "location": "Accra",
            "region": "Greater Accra",
            "bandwidth": 100,
            "service_type": "Dedicated Internet",
            "confirmation_date": "2026-05-02",
            "target_completion_date": "2026-06-10",
            "status": "Implementation",
            "mrc": 3200,
            "nrc": 1200,
            "planned_adss_distance_m": 450,
            "planned_drop_distance_m": 80,
            "poles_9m": 4,
            "remarks": "Civil works in progress",
        },
        {
            "customer_name": "Northstar Bank",
            "site_name": "Kumasi DR Site",
            "location": "Kumasi",
            "region": "Ashanti",
            "bandwidth": 200,
            "service_type": "MPLS",
            "confirmation_date": "2026-04-18",
            "target_completion_date": "2026-06-03",
            "completion_date": "2026-06-15",
            "status": "Delivered",
            "mrc": 5800,
            "nrc": 2400,
            "planned_adss_distance_m": 900,
            "planned_drop_distance_m": 120,
            "poles_11m": 6,
        },
    ]
    for sample in samples:
        save_project(sample)


def main() -> None:
    if REQUIRE_DATABASE_URL and not DATABASE_URL:
        raise RuntimeError("DATABASE_URL is required for this deployment")
    init_db()
    seed_if_empty()
    port = int(os.environ.get("PORT", "8765"))
    host = os.environ.get("HOST", "127.0.0.1")
    server = ThreadingHTTPServer((host, port), TrackerHandler)
    print(f"Project Tracker running at http://{host}:{port}")
    server.serve_forever()


if __name__ == "__main__":
    main()
